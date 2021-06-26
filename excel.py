import openpyxl
from openpyxl.utils import get_column_letter


def get_values_columns(path_workbook, name_worksheet, column1, column2, beginning_row, ending_row):
    '''
    Get values of all cells in the two given columns. If cell value is not an integer, also place None into the list,
    so that both lists have the same length. path_workbook e.g.
    '''

    list_values1 = []
    list_values2 = []

    wb = openpyxl.load_workbook(path_workbook)
    ws = wb[name_worksheet]

    for cell in ws[column1]:
        if beginning_row <= cell.row <= ending_row:
            list_values1.append(cell.value)

    for cell in ws[column2]:
        if beginning_row <= cell.row <= ending_row:
            list_values2.append(cell.value)

    return list_values1, list_values2

def insert_values_column(path_workbook, name_worksheet, column, list_values_to_be_inserted, beginning_row):
    wb = openpyxl.load_workbook(path_workbook)
    ws = wb[name_worksheet]

    for index in range(len(list_values_to_be_inserted)):
        # e.g. ws['A1'] = 'fresh_avocado'
        ws[column + str(index + beginning_row)] = list_values_to_be_inserted[index]

    wb.save(path_workbook)
